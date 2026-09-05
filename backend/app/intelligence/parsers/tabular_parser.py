"""Tabular & Structured Data Parser for CSV, TSV, JSON, XML, and Excel spreadsheets."""

import csv
import io
import json
import os
import defusedxml.ElementTree as ET
from defusedxml.common import DefusedXmlException
from typing import List, Optional
import openpyxl

from app.intelligence.models import (
    Document,
    DocumentElement,
    ElementType,
    TableData,
)
from app.intelligence.parsers.base import BaseParser, CorruptedDocumentError
from app.intelligence.parsers.decoder import read_text_file_strictly

TABULAR_PARSER_VERSION = "1.1.0"
MAX_ROWS_PER_TABLE_CHUNK = 50


class TabularParser(BaseParser):
    """
    Parser for structured datasets: CSV, TSV, JSON, XML, and Microsoft Excel (.xlsx, .xls).
    Extracts tabular headers, row data, and worksheet hierarchies with bounded row-grouping.
    """

    @property
    def parser_name(self) -> str:
        return "tabular-parser"

    @property
    def parser_version(self) -> str:
        return TABULAR_PARSER_VERSION

    @property
    def supported_mime_types(self) -> List[str]:
        return [
            "text/csv",
            "text/tab-separated-values",
            "application/json",
            "application/xml",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ]

    @property
    def supported_extensions(self) -> List[str]:
        return [".csv", ".tsv", ".json", ".xml", ".xlsx", ".xls"]

    def parse(self, file_path: str, file_id: str, mime_type: str = "text/csv") -> Document:
        file_path = str(file_path)
        filename = os.path.basename(file_path)
        ext = os.path.splitext(filename)[1].lower()

        doc_obj = Document(
            file_id=file_id,
            source_path=file_path,
            filename=filename,
            mime_type=mime_type,
            title=filename,
            parser_name=self.parser_name,
            parser_version=self.parser_version,
            total_pages=1,
        )

        try:
            if ext in (".csv", ".tsv"):
                self._parse_delimited(file_path, doc_obj, file_id, ext)
            elif ext == ".json":
                self._parse_json(file_path, doc_obj, file_id)
            elif ext == ".xml":
                self._parse_xml(file_path, doc_obj, file_id)
            elif ext == ".xlsx":
                self._parse_xlsx(file_path, doc_obj, file_id)
            elif ext == ".xls":
                self._parse_legacy_xls(file_path, doc_obj, file_id)
        except Exception as exc:
            if isinstance(exc, CorruptedDocumentError):
                raise
            raise CorruptedDocumentError(f"Failed to parse tabular file {filename}: {str(exc)}") from exc

        return doc_obj

    def _parse_delimited(self, file_path: str, doc: Document, file_id: str, ext: str):
        content = read_text_file_strictly(file_path, filename=doc.filename)
        
        # Sniff delimiter
        delimiter = "\t" if ext == ".tsv" else ","
        if ext == ".csv" and content:
            sample = content[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ","

        reader = csv.reader(io.StringIO(content), delimiter=delimiter)
        rows = [row for row in reader if any(cell.strip() for cell in row)]

        if not rows:
            doc.elements.append(
                DocumentElement(
                    element_id=f"{file_id}_elem_1",
                    element_type=ElementType.PARAGRAPH,
                    text=f"Empty tabular file: {doc.filename}",
                    media_type="tabular",
                    extraction_method="native",
                )
            )
            return

        headers = [c.strip() for c in rows[0]]
        data_rows = [[c.strip() for c in r] for r in rows[1:]]

        element_idx = 0

        if not data_rows:
            # Only headers
            t_data = TableData(headers=headers, rows=[], caption=doc.filename)
            element_idx += 1
            doc.elements.append(
                DocumentElement(
                    element_id=f"{file_id}_elem_{element_idx}",
                    element_type=ElementType.TABLE,
                    text=t_data.to_markdown(),
                    table_data=t_data,
                    sheet_name=os.path.splitext(doc.filename)[0],
                    line_start=1,
                    line_end=1,
                    media_type="tabular",
                    extraction_method="native",
                )
            )
            return

        # Bounded row-group chunking (preserving header on each chunk)
        for chunk_start in range(0, len(data_rows), MAX_ROWS_PER_TABLE_CHUNK):
            chunk_slice = data_rows[chunk_start:chunk_start + MAX_ROWS_PER_TABLE_CHUNK]
            start_row_num = chunk_start + 2  # 1-indexed accounting for header
            end_row_num = start_row_num + len(chunk_slice) - 1
            
            caption = f"{doc.filename} (Rows {start_row_num}-{end_row_num})"
            t_data = TableData(headers=headers, rows=chunk_slice, caption=caption)
            element_idx += 1
            doc.elements.append(
                DocumentElement(
                    element_id=f"{file_id}_elem_{element_idx}",
                    element_type=ElementType.TABLE,
                    text=t_data.to_markdown(),
                    table_data=t_data,
                    sheet_name=os.path.splitext(doc.filename)[0],
                    line_start=1 if chunk_start == 0 else start_row_num,
                    line_end=end_row_num,
                    media_type="tabular",
                    extraction_method="native",
                    metadata={"row_start": start_row_num, "row_end": end_row_num, "total_rows": len(data_rows)},
                )
            )

    def _parse_json(self, file_path: str, doc: Document, file_id: str):
        content = read_text_file_strictly(file_path, filename=doc.filename)
        data = json.loads(content)

        element_idx = 0
        if isinstance(data, list) and data and isinstance(data[0], dict):
            # Array of objects -> Table
            headers = list(data[0].keys())
            rows = [[str(item.get(k, "")) for k in headers] for item in data]
            
            elem_h = DocumentElement(
                element_id=f"{file_id}_elem_1",
                element_type=ElementType.HEADING,
                text=f"Dataset: {doc.filename} ({len(rows)} entries)",
                level=1,
                media_type="tabular",
                extraction_method="native",
            )
            doc.elements.append(elem_h)
            element_idx = 1

            for chunk_start in range(0, len(rows), MAX_ROWS_PER_TABLE_CHUNK):
                chunk_slice = rows[chunk_start:chunk_start + MAX_ROWS_PER_TABLE_CHUNK]
                start_row_num = chunk_start + 1
                end_row_num = start_row_num + len(chunk_slice) - 1
                caption = f"{doc.filename} (Items {start_row_num}-{end_row_num})"
                t_data = TableData(headers=headers, rows=chunk_slice, caption=caption)
                element_idx += 1
                doc.elements.append(
                    DocumentElement(
                        element_id=f"{file_id}_elem_{element_idx}",
                        element_type=ElementType.TABLE,
                        text=t_data.to_markdown(),
                        table_data=t_data,
                        parent_heading_id=elem_h.element_id,
                        media_type="tabular",
                        extraction_method="native",
                        metadata={"item_start": start_row_num, "item_end": end_row_num},
                    )
                )
        elif isinstance(data, dict):
            elem_h = DocumentElement(
                element_id=f"{file_id}_elem_1",
                element_type=ElementType.HEADING,
                text=f"Document: {doc.filename}",
                level=1,
                media_type="tabular",
                extraction_method="native",
            )
            doc.elements.append(elem_h)
            element_idx = 1

            for key, val in data.items():
                element_idx += 1
                h_sec = DocumentElement(
                    element_id=f"{file_id}_elem_{element_idx}",
                    element_type=ElementType.HEADING,
                    text=f"Property: {key}",
                    level=2,
                    parent_heading_id=elem_h.element_id,
                    media_type="tabular",
                    extraction_method="native",
                )
                doc.elements.append(h_sec)

                element_idx += 1
                val_str = json.dumps(val, indent=2) if isinstance(val, (dict, list)) else str(val)
                doc.elements.append(
                    DocumentElement(
                        element_id=f"{file_id}_elem_{element_idx}",
                        element_type=ElementType.PARAGRAPH,
                        text=val_str,
                        parent_heading_id=h_sec.element_id,
                        media_type="tabular",
                        extraction_method="native",
                    )
                )
        else:
            doc.elements.append(
                DocumentElement(
                    element_id=f"{file_id}_elem_1",
                    element_type=ElementType.PARAGRAPH,
                    text=json.dumps(data, indent=2),
                    media_type="tabular",
                    extraction_method="native",
                )
            )

    def _parse_xml(self, file_path: str, doc: Document, file_id: str):
        content = read_text_file_strictly(str(file_path), filename=doc.filename)
        try:
            root = ET.fromstring(content)
        except Exception as exc:
            raise CorruptedDocumentError(f"Malformed or unsafe XML in {doc.filename}: {exc}") from exc

        element_idx = 1
        root_tag = root.tag.split("}")[-1] if "}" in root.tag else root.tag
        root_attrs = " ".join(f'{k}="{v}"' for k, v in root.attrib.items()) if root.attrib else ""
        elem_h = DocumentElement(
            element_id=f"{file_id}_elem_1",
            element_type=ElementType.HEADING,
            text=f"XML Document: <{root_tag}{(' ' + root_attrs) if root_attrs else ''}>",
            level=1,
            media_type="tabular",
            extraction_method="native",
        )
        doc.elements.append(elem_h)

        def _traverse(node, parent_id, depth=2):
            nonlocal element_idx
            node_tag = node.tag.split("}")[-1] if "}" in node.tag else node.tag
            text = (node.text or "").strip()
            attrs_str = " ".join(f'{k}="{v}"' for k, v in node.attrib.items()) if node.attrib else ""
            
            if text:
                element_idx += 1
                attr_part = f" ({attrs_str})" if attrs_str else ""
                doc.elements.append(
                    DocumentElement(
                        element_id=f"{file_id}_elem_{element_idx}",
                        element_type=ElementType.PARAGRAPH,
                        text=f"<{node_tag}>{attr_part}: {text}",
                        parent_heading_id=parent_id,
                        media_type="tabular",
                        extraction_method="native",
                    )
                )
            elif attrs_str and len(node) == 0:
                element_idx += 1
                doc.elements.append(
                    DocumentElement(
                        element_id=f"{file_id}_elem_{element_idx}",
                        element_type=ElementType.PARAGRAPH,
                        text=f"<{node_tag} {attrs_str}/>",
                        parent_heading_id=parent_id,
                        media_type="tabular",
                        extraction_method="native",
                    )
                )

            for child in node:
                child_tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag
                child_attrs = " ".join(f'{k}="{v}"' for k, v in child.attrib.items()) if child.attrib else ""
                if len(child) > 0:
                    element_idx += 1
                    ch_h = DocumentElement(
                        element_id=f"{file_id}_elem_{element_idx}",
                        element_type=ElementType.HEADING,
                        text=f"Section: <{child_tag}{(' ' + child_attrs) if child_attrs else ''}>",
                        level=min(depth, 4),
                        parent_heading_id=parent_id,
                        media_type="tabular",
                        extraction_method="native",
                    )
                    doc.elements.append(ch_h)
                    _traverse(child, ch_h.element_id, depth + 1)
                else:
                    _traverse(child, parent_id, depth)

        _traverse(root, elem_h.element_id)

    def _parse_xlsx(self, file_path: str, doc: Document, file_id: str):
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        try:
            doc.total_pages = len(wb.sheetnames)
            element_idx = 0

            for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
                sheet = wb[sheet_name]
                rows_data = []
                for row in sheet.iter_rows(values_only=True):
                    if any(c is not None and str(c).strip() for c in row):
                        rows_data.append([str(c or "").strip() for c in row])

                if rows_data:
                    element_idx += 1
                    elem_h = DocumentElement(
                        element_id=f"{file_id}_elem_{element_idx}",
                        element_type=ElementType.HEADING,
                        text=f"Worksheet: {sheet_name}",
                        page_number=sheet_idx,
                        sheet_name=sheet_name,
                        level=1,
                        media_type="tabular",
                        extraction_method="native",
                    )
                    doc.elements.append(elem_h)

                    headers = rows_data[0]
                    body_rows = rows_data[1:] if len(rows_data) > 1 else []

                    if not body_rows:
                        t_data = TableData(headers=headers, rows=[], caption=sheet_name)
                        element_idx += 1
                        doc.elements.append(
                            DocumentElement(
                                element_id=f"{file_id}_elem_{element_idx}",
                                element_type=ElementType.TABLE,
                                text=t_data.to_markdown(),
                                page_number=sheet_idx,
                                sheet_name=sheet_name,
                                table_data=t_data,
                                parent_heading_id=elem_h.element_id,
                                line_start=1,
                                line_end=1,
                                media_type="tabular",
                                extraction_method="native",
                            )
                        )
                    else:
                        # Bounded row-group chunking for spreadsheets
                        for chunk_start in range(0, len(body_rows), MAX_ROWS_PER_TABLE_CHUNK):
                            chunk_slice = body_rows[chunk_start:chunk_start + MAX_ROWS_PER_TABLE_CHUNK]
                            start_row_num = chunk_start + 2
                            end_row_num = start_row_num + len(chunk_slice) - 1
                            caption = f"{sheet_name} (Rows {start_row_num}-{end_row_num})"
                            t_data = TableData(headers=headers, rows=chunk_slice, caption=caption)
                            element_idx += 1
                            doc.elements.append(
                                DocumentElement(
                                    element_id=f"{file_id}_elem_{element_idx}",
                                    element_type=ElementType.TABLE,
                                    text=t_data.to_markdown(),
                                    page_number=sheet_idx,
                                    sheet_name=sheet_name,
                                    table_data=t_data,
                                    parent_heading_id=elem_h.element_id,
                                    line_start=1 if chunk_start == 0 else start_row_num,
                                    line_end=end_row_num,
                                    media_type="tabular",
                                    extraction_method="native",
                                    metadata={"sheet_name": sheet_name, "row_start": start_row_num, "row_end": end_row_num},
                                )
                            )
        finally:
            wb.close()

    def _parse_legacy_xls(self, file_path: str, doc: Document, file_id: str):
        """Extracts text runs and tabular records from legacy binary .xls files."""
        from app.intelligence.parsers.legacy_doc_ppt_parser import extract_binary_text_streams
        runs = extract_binary_text_streams(file_path)
        
        elem_idx = 1
        elem_h = DocumentElement(
            element_id=f"{file_id}_elem_1",
            element_type=ElementType.HEADING,
            text=f"Spreadsheet: {doc.filename} (XLS)",
            level=1,
            media_type="tabular",
            extraction_method="native",
        )
        doc.elements.append(elem_h)

        for r in runs:
            elem_idx += 1
            doc.elements.append(
                DocumentElement(
                    element_id=f"{file_id}_elem_{elem_idx}",
                    element_type=ElementType.PARAGRAPH,
                    text=r,
                    parent_heading_id=elem_h.element_id,
                    media_type="tabular",
                    extraction_method="native",
                )
            )
